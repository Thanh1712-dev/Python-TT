#!/usr/bin/env python
# coding: utf-8

#  有时候，我们需要从一长串字符串中提取中文，比如如下这样的：
#  ![](images\problem.png)
#  我们可以看到，中文的长度参差不齐，在字符串中的位置也不固定。因此无论是用Excel自带的`left`,`right`,`mid` 函数，还是使用分列都无能为力。下面介绍通过Python的正则表达式，一键轻松提取中文。

# In[2]:


from IPython.core.interactiveshell import InteractiveShell
InteractiveShell.ast_node_interactivity = "all"


# In[3]:


#提取数据，以便后续处理
from openpyxl import load_workbook 
data=[]
wb = load_workbook('data/data.xlsx')
ws = wb.active
for row in range(2,ws.max_row+1): #从第二行开始遍历excel文件所有行
    info = ws['A' + str(row)].value
    data.append(info)
data[:5]
len(data)


# 以上，我们先打开包含待处理的Excel文档，然后将待处理的字符串全部提取，存入列表`data`。然后通过`data[:5]`查看一下前5条数据，确保跟源文件相符。并通过`len(data)`查看数据的条数，也是与原文件是相符的。

# In[5]:


#获取中文
import re #re是正则表达式模块
chinese_list=[]
for i in data:
    chinese = re.findall('[\u4e00-\u9fa5]',i) #汉字的范围为"\u4e00-\u9fa5"
    chinese_list.append(''.join(chinese))
chinese_list[0]


# In[6]:


chinese


# 正则表达式（regular expressions），也可以叫规则表达式，就是通过设定一些规则，从杂乱无章的字符串中捞取我们想要的数据，比如电话号码，邮件地址，邮政编码等。以上，先导入`re`模块，正则表达式的所有函数都在这个模块里面。然后新建一个列表`chinese_list`，以便存储抓取好的数据。这里，我们使用`findall()`在汉字的范围`\u4e00-\u9fa5`内抓取一个字符串中的所有汉字，这个方法返回的一组字符串中会包含被查找字符串中的所有匹配。也就是说只要是汉字，会逐个全部抓出，并将汉字以单个的形式存入列表，比如上面的`['集', '成', '天', '线']`。但是我们希望这四个字是连在一起的，所以在放入`chinese_list`列表之前，先用`join`方法将这几个字连接在一起了。`join`前面的引号内指定用什么符号连接，这里什么符号都没使用，即代表直接连接。下面我们展示一下用“&”符号连接一下，看看效果，其结果是`'集&成&天&线'`。

# In[7]:


"&".join(chinese)


# 连接好汉字后，我们将其写入原Excel表格第2列中，以便于第一列对应，代码如下：

# In[5]:


for row in range(2,ws.max_row+1):
    ws['B' + str(row)].value = chinese_list[row-2]
wb.save('data/中文.xlsx')


# 最后看看结果吧：
# ![](images\result.png)
