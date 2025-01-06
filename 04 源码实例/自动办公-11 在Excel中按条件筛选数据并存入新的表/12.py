#!/usr/bin/env python
# coding: utf-8

# 老板想要看去年每月领料数量大于1000的数据。手动筛选并复制粘贴出来，需要重复操作12次，实在太麻烦了，还是让Python来做吧。磨刀不误砍柴工，先整理一下思路：
# 1. 读取原表，将数量大于1000的数据所对应的行整行提取（如同在excel表中按数字筛选大于1000的）
# 2. 将提取的数据写入新的Excel表
# ![](images\problem.png)

# In[1]:


#1.获取满足条件的数据
from openpyxl import load_workbook
wb = load_workbook("每月物料表.xlsx")
data = {} #储存所有工作表中满足条件的数据，以工作表名称为键
sheet_names = wb.sheetnames
for sheet_name in sheet_names:
    ws = wb[sheet_name]
    qty_list = []
    #获取G列的数据，并用enumrate给其对应的元素编号
    for row in range(2,ws.max_row+1):
        qty = ws['G'+str(row)].value
        qty_list.append(qty)

    qty_idx = list(enumerate(qty_list)) #用于编号
    
    #判断数据是否大于1000，然后返回大于1000的数据所对应的行数
    row_idx = [] #用于储存数量大于1000所对应的的行号
    for i in range(len(qty_idx)):
        if qty_idx[i][1] > 1000:
            row_idx.append(qty_idx[i][0]+2)

    #获取满足条件的数据        
    data_morethan1K = []
    for i in row_idx:
        data_morethan1K.append(ws['A'+str(i)+":"+'I'+str(i)])
        
    data[sheet_name]=data_morethan1K 


# 以上，我们把满足条件的12个月的数据提取并存入字典`data`，其键为对应的月份，比如“1月”，值就是满足条件的各行的数据。我们把“每月物料表”的G列对应的数据提取，存入列表`qty_list`，其中前10个数据是如下这样的。

# In[2]:


qty_list[:10]


# 然后需要使用`enumerate`函数给这个列表的数据加上索引，以便在跟1000比大小的时候定位满足条件的那些数据的对应在Excel表中的行数。加上索引之后的列表是如下这样的，索引从0开始累加。

# In[3]:


qty_idx[:10]


# 然后，再新建一个列表`row_idx`，用于储存“领料数量”大于1000的数据所对应的行号。此处用到`if`语句进行判断，只将“领料数量”大于1000的数据所对应的行号加上2存入列表。为什么要加2，是因为`range`函数是从0开始取的，然后工作表首行是字段名，第二行开始才是数据。如下结果显示了满足条件的数据对应的行数。

# In[4]:


row_idx[:5]


#  然后新建列表`data_morethan1K`用于存储以上行号对应的整行数据。比如`ws['A1:I1']`就指第一行从A列到I列的所有单元格数据。最后将数据存入`data`字典中。数据结构如下所示。

# In[5]:


data_morethan1K[1]


# In[6]:


data['1月']


# In[7]:


len(data['1月'])


# In[8]:


data['1月'][0][0][1].value


# 数据提取完成后，就可以开始写入数据了。打开模板，按月从`data`字典中提取数据。并根据数据结构找到层级关系，将其中的各行的数据写入各单元格。写完之后，设置一下字号、边框即对齐方式，保存数据。到此收工！

# In[9]:


#2.写入获取的数据
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
thin = Side(border_style="thin", color="000000")#定义边框粗细及颜色

wb = load_workbook("模板.xlsx")
ws = wb.active
for month in data.keys():
    ws_new = wb.copy_worksheet(ws) #复制模板中的工作表
    ws_new.title=month 
    #将每个月的数据条数逐个取出并写入新的工作表
    for i in range(len(data[month])): #按数据行数计数，每行数据对应9列，所以每行需分别写入9个单元格
        ws_new.cell(row=i+2,column=1).value=data[month][i][0][0].value
        ws_new.cell(row=i+2,column=2).value=data[month][i][0][1].value
        ws_new.cell(row=i+2,column=3).value=data[month][i][0][2].value
        ws_new.cell(row=i+2,column=4).value=data[month][i][0][3].value.date()
        ws_new.cell(row=i+2,column=5).value=data[month][i][0][4].value
        ws_new.cell(row=i+2,column=6).value=data[month][i][0][5].value
        ws_new.cell(row=i+2,column=7).value=data[month][i][0][6].value
        ws_new.cell(row=i+2,column=8).value=data[month][i][0][7].value
        ws_new.cell(row=i+2,column=9).value=data[month][i][0][8].value
    
    #设置字号，对齐，缩小字体填充，加边框
    #Font(bold=True)可加粗字体

    for row_number in range(2, ws_new.max_row+1):
        for col_number in range(1,10):
            c = ws_new.cell(row=row_number,column=col_number)
            c.font = Font(size=10)
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            c.alignment = Alignment(horizontal="left", vertical="center",shrink_to_fit = True)
wb.save("每月(大于1K).xlsx")


#  华丽的结果如下：
#  ![](images\result.png)
