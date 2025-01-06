#!/usr/bin/env python
# coding: utf-8

# 系统记录了每个部门的领料情况，现在要求汇总所有部门的领料明细，计算并加总各部门每日领料次数。各部门领料明细如下，需要抓取的数据在红色虚线框里。但是“业务类型”字段中的“备注”数据不需要。

# ![](images\requirement.png)

# 汇总后的数据要求在原表基础上新增5列，数据来源都是原表。其中领取日期、时间需要在原表的“时间”字段中截取。领取次数要求加总，比如2018年1月13日这天领了三次料，则这三次领料记录对应的“领取次数”都填上“3”。“领料明细汇总表”的文件名按“领料明细汇总表”加上日期、时间命名，比如“领料明细汇总表2020-06-24 10.30.11.xlsx”。

# ![](images\result0.png)

# 共分两大步走。第一步，提取并处理好数据。第二步，新建Excel文件，写入数据，调整格式并保存。
# <br/>
# <br/>先看第一步。观察文件，发现系统导出的都是'.xls'的文件，因此需要用支持该文件库来处理。`openpyxl`只能歇菜了，这家伙只能搞".xlsx"文件。`xlwings`和`xlrd`可以考虑。用`xlwings`试了试，因为文件是“受保护的视图”，根本无法读取，也靠边站着凉快去了。只有`xlrd`不负众望，轻松读取。就你了，`xlrd`。什么？`import`失败？先运行`cmd`，输入`"pip install xlrd"`安装好了再来。
# ![](images\error.png)

# 如下定义了一个函数`Get_data()`，其中有一个参数`file`，即被读取的文件路径。先用`xlrd.open_workbook`打开工作簿，然后选第一个工作表（工作簿中只有一个工作表），存入`ws`。建一个字典`data`用于存数据。然后用`for`循环遍历领料明细表对应行列，提取数据。
# <br/>
# <br/>提取到时间这里，出现了问题。原来原表中，有些时间是浮点型，有些是字符串型，一刀切式的处理就出问题了。看来得架两个锅，分别烹饪了。于是上`if`语句，两种不同类型的分别处理。`xlrd`从Excel文件单元格中读取的时间数据是浮点型，比如“43460.299733796295”。需要使用`xlrd.xldate.xldate_as_datetime(xldate, datemode)`将其转换为Python的标准时间格式。其中第一个参数`xldate`就是那个从Excel文件中读取的浮点型数据，第二个参数`datemode`有“0”和“1”两个值，指基于1900还是1904，一般我们是基于1900，所以选“0”。转换后的时间如下，对人类很友好。

# In[15]:

import xlrd
xlrd.xldate.xldate_as_datetime(43460.299733796295,0)


# 而对于字符串格式的，是这样的`'2019-12-25 09:04:10'`，以上函数无法识别，`datetime.strptime()`善于搞这个。

# In[16]:

import datetime
datetime.datetime.strptime('2019-12-25 09:04:10','%Y-%m-%d %H:%M:%S')


# 依次逐个正常提取数据，按下不表，直到“领料日期”和“领料时间”。这二位不在原表中，需要从上面处理好的`date_time`中截取，分别使用了`date_time.date()`和`date_time.time()`。以上的时间数据就被分割成了`datetime.date(2019, 12, 25)`和`datetime.time(9, 4, 10)`。最后将所有数据规规整整，按照顺序存入列表`info_list`中。然后将日期作为字典`data`的键的默认值，将相同日期的领料数据作为值存入列表。由于我们不需要“业务类型”为“备注”的数据，所以加了一个`if`判断语句，将其排除在外。
# <br/>
# <br/>到此，数据还没完全OK，因为当天的领料次数还未计算。用`for`循环，遍历字典`data`，加总每天领料的数据条数，并将其插入到每条数据的最后一个位置。最后返回最终的字典`data`，至此，一张Excel表中的数据算是搞定了。
# <br/>
# <br/>可能有小伙伴会问，提取数据的时候为什么不整行取呢？有道理，我也想啊，哪料那破系统导出的数据有很多合并单元格，不连续，整行提取出来还要整理一番，太麻烦。

# In[9]:


#读取xls文件中的数据
import xlrd
import datetime
def Get_data(file):
    wb = xlrd.open_workbook(file) #读取工作簿
    ws = wb.sheets()[0] #选第一个工作表
    data = {}

    for row in range(7, ws.nrows-2):
        dept = ws.cell(2, 16).value #部门
        dept_id = ws.cell(3, 16).value #部门编号
        dt = ws.cell(row, 0).value #时间
        if type(dt) is float:
            date_time = xlrd.xldate.xldate_as_datetime(dt, 0)
        else:
            date_time = datetime.datetime.strptime(dt,'%Y-%m-%d %H:%M:%S')
        business = ws.cell(row, 2).value #业务类型
        model = ws.cell(row, 3).value #品种
        qty = ws.cell(row, 4).value #数量
        unit_price = ws.cell(row, 6).value #单价
        price = ws.cell(row, 8).value #总价
        reward = ws.cell(row, 9).value #额外值
        discount = ws.cell(row, 11).value #调整
        balance = ws.cell(row, 13).value #剩余
        location = str(ws.cell(row, 15).value).strip() #库位
        operator = ws.cell(row, 17).value #操作员
        date = date_time.date() #日期
        time = date_time.time() #时间
        info_list=[dept,dept_id,date_time,business,model,qty,unit_price,price,reward,discount,
                   balance,location,operator,date,time]
        data.setdefault(date,[]) #以日期为键
        if info_list[3] != "备注": #不要业务类型为“备注”的数据
            data[date].append(info_list)
            
    #增加当日领取次数        
    for key in data.keys():
        for i in data[key]:
            i.append(len(data[key]))

    return data


# 然后，将目标Excel文件的路径全部获取，并存到列表`files`。以便后续使用程序逐个获取。

# In[17]:


import os #用于获取目标文件所在路径
path=os.getcwd()+"\\记录\\" # 文件夹绝对路径
files=[]
for file in os.listdir(path):
    if file.endswith(".xls"): #只获取".xls"后缀的文件
        files.append(path+file) 
files


# 因为汇总后的Excel文件需要用当前日期和时间命名，所以再定义一个函数`Get_current_time`获取当前时间。调用一下，就获得我们设定格式的日期时间了。

# In[18]:


import time
def Get_current_time():
    time_stamp = time.time()  # 当前时间的时间戳
    local_time = time.localtime(time_stamp)  #
    str_time = time.strftime('%Y-%m-%d %H.%M.%S', local_time)
    return str_time
Get_current_time()


# 下面开始第二大步，读取所有目标Excel文件的数据，写入汇总表，并设置相关单元格的格式，隐藏对应的列。`openpyxl`坐了一会儿冷板凳，现在终于可以出场了，忍不住理了理帅气的头发。
# <br/>
# <br/>先导入各种相关的库。然后将字段名存入`title`列表，以便在写入Excel文件的时候使用`append()`整行写入。然后新建Excel工作簿，取其活动工作表，存入`ws`变量。将首行的单元格全部合并，标题行中写入字符串“领料明细汇总表”，设定其字体，加粗和字号，行高，对齐方式。然后写入字段行，一个`append()`就将列表`title`中的元素逐个取出，每个单元格写入一个，非常强大。
# <br/>
# <br/>现在来到写入各部门领料数据的部分。这部分处理数据最多，但程序非常少，那就是定义函数的好处。只需调用函数`Get_data`，然后将每个领料明细表中的数据取出整理好，然后使用`append()`整行写入即可。
# <br/>
# <br/>数据写完后，将整个数据部分的单元格格式设置成想要的。然后再设置一下列宽，避免部分字段的数据太长而无法完全显示。这里，我们将列名和列宽的值都存在列表中，然后逐个进行设置，这比单个地设置简洁很多。简洁是Python的特色之一，必须强烈支持。最后分组隐藏设定的列，将当前日期时间加入文件名，保存数据。

# In[20]:


#汇总数据到主文件
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment #设置单元格格式
thin = Side(border_style="thin", color="000000")#定义边框粗细及颜色
title = ['部门', '部门编号', '时间', '业务类型', '品种', '数量', '单价', '金额', '额外值',
         '调整', '剩余', '库位', '操作员', '领取日期', '领取时间', '领取次数']

wb = Workbook() 
ws = wb.active
ws.merge_cells("A1:P1") #合并首行单元格
ws.cell(1,1).value = "领料明细汇总表"
ws.cell(1,1).font = Font(name=u'黑体',bold=True,size=18)
ws.row_dimensions[1].height  = 22.2 #设置首行行高
ws.cell(1,1).alignment = Alignment(horizontal="center", vertical="center") #设置对齐
ws.append(title) #写入字段行

#写入各部门领料的数据
for file in files:
    data = Get_data(file)
    for key in data.keys():
        for i in data[key]:
            ws.append(i)

#设置字号，对齐，缩小字体填充，加边框
#Font(bold=True)可加粗字体
for row_number in range(2, ws.max_row+1):
    for col_number in range(1,17):
        c = ws.cell(row=row_number,column=col_number)
        c.font = Font(size=9)
        c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        c.alignment = Alignment(horizontal="left", vertical="center")
        
#设置列宽        
col_name= list("ABCDEFGHIJKLMNOP")
col_width = [8, 8, 16, 8, 16, 8, 8, 9.8, 8, 8, 8, 11, 8.3, 9, 8, 8]
for i in range(len(col_name)):
    ws.column_dimensions[col_name[i]].width = col_width[i]
    
#分组隐藏列
ws.column_dimensions.group('I','K',hidden=True)
ws.column_dimensions.group('N','O',hidden=True)

wb.save(f"领料明细汇总表{Get_current_time()}.xlsx")


# 上结果图。相当美丽，想说不爱你，铿锵有力的一个字：难！
# ![](images\result1.png)
