#!/usr/bin/env python
# coding: utf-8

# 随着项目的进展，需要经常在Excel业务表格中查找及替换数据，已保证数据与实际项目进度一致。手动一个一个查找，然后替换，效率太低，还容易遗漏。现在我们来试试用Python自动完成查找及替换吧。具体要求如下。
# ![](images\requirement.png)

# 首先，我们先将左边表格中的数据提取，并存入字典`data`，其键为“查找内容”中的数据，值为“替换内容”中的数据。

# In[1]:


from openpyxl import load_workbook #用于读取Excel中的信息
#获取Excel表格中的数据
wb = load_workbook('查找替换.xlsx')#读取工作簿
ws = wb.active #读取活动工作表
data={} #新建字典，用于储存数据

for row in range(2,ws.max_row+1):
    chazhao = str(ws['A' + str(row)].value)  #转换成字符串，以免后续比对时出现数据类型冲突
    tihuan = str(ws['B' + str(row)].value) #转换成字符串，以免后续比对时出现数据类型冲突
    data[chazhao]=tihuan #键值对应存入字典


# In[ ]:


data


# 然后，读取目标表格，将D列中的所有数据提取出来，以便后续比对及替换。通过`for`循环遍历“原表”，将D列每个单元格的值提取并存入`ID_list`。然后通过切片`ID_list[:10]`查看前10个数据是否OK。结果显示相当正常。

# In[ ]:


wb = load_workbook('原表.xlsx') #读取目标工作簿
ws = wb.active
ID_list = [] #新建一个列表，用于储存原表D列的信息
for row in range(2,ws.max_row+1):
    ID = ws['D' + str(row)].value #遍历整个工作表，将D列的数据逐个存入ID变量
    ID_list.append(ID) #将读取到的结果存入列表
ID_list[:10] #查看列表中前10个数据


# In[ ]:


type("")


# 为了比对数据，我们需要将`'说明码：77601'`中的“说明码：”字符拿掉，只保留“77601”。于是调用`split`函数来进行分割，并将分割好的数字部分存入新建的列表`code`。不好，居然报错了，说`ID_list`列表中有"None"（空）类型的数据，而"None"类型的数据是不能使用`split`函数的。目测了一下，`ID_list`列表中除了`'说明码：77601'`和`''`这样的空字符串，没看到None啊。再回来“原表”侦察一下，发现最下面还有一些单元格很有嫌疑。原来是表尾有一些“供应商”和仓位信息，这些信息所在位置对应的D列都是空单元格，其值为"None"。用`ID_list[-10:]`查看最后10个元素，果然都是"None"。
# ![](images\error.png)

# In[ ]:


# code = [i.split("：")[-1] for i in ID_list]
# code


# In[ ]:


ID_list[-10:]


# 这样，我们就知道`ID_list`中有三种数据，即含内容的字符串（比如'说明码：77601'），空字符串（比如''）和空值None。因此，需要修改一下字符串分割代码如下。加入了`if`判断语句，如果`ID_list`中的值是None，则放入None占位，以保持列表的值的顺序与原表一致；值不是None，则按"："符号分割，并放分割后的最后一个值`[-1]`进入新列表code。空字符串在这里也要经过`split`分割，但其中没有“：”符号，所以就分割不了，只得直接跳过，最后放入新列表的还是空字符串。

# In[ ]:


code = []
for i in ID_list:
    if i == None:# 如果是None，则放入None占位，以保持列表的值的顺序与原表一致
        code.append(None)
    else:
        code.append(i.split("：")[-1]) #不是None,则按"："符号分割，并放分割后的最后一个值进入新列表code
code[:10]


# 处理完数据后，即可开始查找并替换目标数据了。用`for`循环遍历列表`code`，即原表D列中的数字部分。如果其中的值也存在于data的键中，即语句`if code[i] in data`，则将原表中D列(`column=4`)对应的行中的数据改写成新的值。新的值由两部分组成，一部分是“说明码：”这样的，即`ID_list[i].split("：")[0]`，另一部分则是要替换的数字，即`data[code[i]]`。这样保证只替换了需要替换的数字部分，而保留中文和冒号部分。最后保存为新的文件，替换完成。

# In[ ]:


for i in range(len(code)):
    if code[i] in data:
        ws.cell(row=i+2,column=4).value = ID_list[i].split("：")[0] +"："+ data[code[i]]
wb.save('原表-替换.xlsx') 


# 如果以上不能通过观察原表，发现空值问题，还可以用`enumerate`函数给列表里的所有元素加上索引，以便精确定位`ID_list`中的空值。加上索引后，在转换成列表，并存入新的列表`ID_list_idx`中。观察其中前10个数据，可见索引已加好了。然后遍历新列表，判断其中的值是否为空值，若是则打印其对应的索引编号，这样就能精准定位哪些是空值了，再回到原Excel表，就容易弄清楚发生了什么事啦。其中，新列表中的元素的结构是一个元组，像这样`(2, '说明码：77601')`，`i[0]`是索引`2`，`i[1]`是索引对应的值`说明码：77601`。

# In[ ]:


ID_list_idx = list(enumerate(ID_list)) #加索引
ID_list_idx[:10]


# In[3]:


for i in ID_list_idx: #遍历列表
    if i[1] == None: #判断索引对应的值是否为空值。i的结构是一个元组，像这样(2, '说明码：77601')，i[0]是索引，i[1]是索引对应的值
        print(i[0]) #打印索引编号

