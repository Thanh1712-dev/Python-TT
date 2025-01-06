#!/usr/bin/env python
# coding: utf-8

# “物料表”中包含全年12个月的领用记录，现在要求将5月和8月的记录筛选出来新建工作表存放，并从原表中将这部分数据删掉。
# ![](images\requirement.png)
# 我是骚气的示意图。

# 于是乎，开始写码。先导入`pandas`库，然后指定标题行(header = 2)，读取Excel工作簿。展示其前5行，整齐划一，没问题。

# In[1]:


import pandas as pd
df = pd.read_excel("物料表.xlsx", header  = 2)
df.head()


# 由于要按月份筛选，而原表格没有月份这个字段，因此新增月份字段。月份从哪里来呢？直接从原表“日期”字段中提取，`.dt.month`即可提取原日期中的月份。再展示前5行，最后面多了1列“月份”。一切顺利，继续。

# In[2]:


df['月份']=df['日期'].dt.month
df.head()


# 下面要删选出5月的数据，于是`df[df['月份']==5]`暴力提取，并存入df5。展示一下前5行，正常。同样的方法删选8月份的数据。

# In[3]:


df5 = df[df['月份']==5]
df5.head()


# In[4]:


df8 = df[df['月份']==8]
df8.head()


# 然后获取除开5月和8月的数据。因涉及到批量重复操作，还是定义个函数比较简洁。函数`Remove_data`包含两个参数，一个是df（即传入的DataFrame），另一个是月份列表。遍历月份列表，逐个剔除所选月份的数据。`df[df['月份']!=i]`表示只保留"i"之外的数据，"!=" 是不等号。此例，先剔除5月的数据，再剔除8月的数据，剔除好的数据存入`df_rest`。

# In[5]:


#去掉不需要的月份的数据
def Remove_data(df,month=[]):
    for i in month:
        df = df[df['月份']!=i]
    return df

df_rest = Remove_data(df,[5,8])
df_rest.head()


# 然后调用`ExcelWriter`将三个数据写入同一个Excel文件。在写入前，先用`drop`将月份列删除，它的使命已经完成，可以光荣退休了。`ExcelWriter`在实例52中有详细说明，请出门右转。在正式使用的时候，Excel文件名直接使用"物料表.xlsx"，则可将原文件覆盖。这里为了展示结果，所以用了文件名`物料表_1.xlsx`来区分。ExcelWriter中的模式使用的是默认“写”模式，因此会覆盖原同名文件。

# In[6]:


with pd.ExcelWriter('物料表_1.xlsx', engine='openpyxl',
                    datetime_format='YYYY-MM-DD') as writer:
    df5.drop('月份',axis = 1).to_excel(writer, sheet_name='5月',index = False)
    df8.drop('月份',axis = 1).to_excel(writer, sheet_name='8月',index = False)
    df_rest.drop('月份',axis = 1).to_excel(writer, sheet_name='剩余月份',index = False)


# 因为`ExcelWriter`写入数据后，是不会调整单元格格式的，这样的结果很难看很不专业，所以需要用如下代码批量操作一下。

# In[7]:


from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
thin = Side(border_style="thin", color="000000")#定义边框粗细及颜色

wb = load_workbook("物料表_1.xlsx")
for sheetname in wb.sheetnames:    
    ws = wb[sheetname]
    # 调整列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['C'].width = 15.5
    ws.column_dimensions['G'].width = 10

    #设置字号，对齐，缩小字体填充，加边框
    for row_number in range(2, ws.max_row+1):
        for col_number in range(1,ws.max_column+1):
            c = ws.cell(row=row_number,column=col_number)
            c.font = Font(size=10)
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            c.alignment = Alignment(horizontal="left", vertical="center")
wb.save("物料表_1.xlsx")


# 上结果图。可见，在“剩余月份”中，5月和8月的数据已经毫不神奇地不见了。就酱。
# ![](images\result.png)
